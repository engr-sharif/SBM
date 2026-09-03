#!/usr/bin/env python3
"""Bake the two site coordinate tables that ship with the app into data/datasets/.

These are ordinary datasets in the js/datasets.js sense — nothing about them is
special-cased in the app — they are simply baked in so the wells and the 2025
geotechnical borings are on the map the moment it opens, instead of having to be
re-imported by hand every session.

Sources (read straight off the project share, nothing inferred from a PDF):
  wells    SBMM\\Groundwater Sampling\\SBMM Monitoring Wells.xlsx, sheet "WellConst"
  borings  SBMM\\Geotechnical\\Sulphur Bank Test Pit Soil Borings Location
           Coordinates20251210.xlsx, sheet "SB"  (coordinates)
           + 2025 Jacobs Investigation\\Borings Schedule and Progress.xlsx (depth
             drilled) + Soil Profiles\\SBMM_Field Interpretted Waste Depth.xlsx

Run:  python tools/build_seed_datasets.py   (needs the staged workbooks)
"""
import json
import os
import re

import openpyxl

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
OUT = os.path.join(ROOT, "data", "datasets")
SRC = "/mnt/user-data/uploads/SBMM"

# EPSG:6418 sanity window for this site. Anything outside is a transcription
# error, not a location — drop it rather than draw it in the wrong county.
XMIN, XMAX = 6_360_000, 6_385_000
YMIN, YMAX = 2_120_000, 2_140_000


def num(v):
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip()
    if s in ("", "M", "NA", "-", "<Null>"):
        return None
    try:
        return float(s)
    except ValueError:
        return None


def txt(v):
    if v is None:
        return None
    if hasattr(v, "date"):
        return v.date().isoformat()
    s = str(v).strip()
    return s if s and s not in ("M", "NA", "-", "<Null>") else None


def in_site(x, y):
    return x and y and XMIN < x < XMAX and YMIN < y < YMAX


# --------------------------------------------------------------------------
def wells():
    p = os.path.join(SRC, "Groundwater Sampling", "SBMM Monitoring Wells.xlsx")
    ws = openpyxl.load_workbook(p, data_only=True)["WellConst"]
    rows = list(ws.iter_rows(values_only=True))
    hdr = [txt(h) for h in rows[0]]
    col = {h: i for i, h in enumerate(hdr) if h}

    def g(r, name):
        i = col.get(name)
        return r[i] if i is not None and i < len(r) else None

    pts, skipped = [], 0
    for r in rows[1:]:
        loc = txt(g(r, "LOC_ID"))
        x, y = num(g(r, "Easting")), num(g(r, "Northing"))
        if not loc or not in_site(x, y):
            skipped += 1
            continue
        a = {
            "Installed": txt(g(r, "Date_Installed")),
            "TOC elev (ft NAVD88)": num(g(r, "Elevation Top of Casing (feet NAVD88)")),
            "Ground elev (ft NAVD88)": num(g(r, "Land Surface (feet NAVD88)")),
            "Casing stickup (ft)": num(g(r, "Casing Stickup (feet)")),
            "Casing diameter (in)": num(g(r, "Casing Diameter (inch)")),
            "Total depth (ft)": num(g(r, "Total Depth (feet)")),
            "Screen top (ft bgs)": num(g(r, "Depth To Top Screen (ft)")),
            "Screen bottom (ft bgs)": num(g(r, "Depth to Bottom Screen (feet)")),
            "Screen elev top (ft)": num(g(r, "Elevation Top of Screen (feet NAVD88)")),
            "Screen elev bottom (ft)": num(g(r, "Elevation Bottom of Screen (feet NAVD88)")),
            "Lithology at screen": txt(g(r, "Lithology at Screen")),
        }
        pts.append({"id": loc, "x": round(x, 2), "y": round(y, 2),
                    "a": {k: v for k, v in a.items() if v is not None}})
    return dict(
        id="wells", name="Monitoring wells", kind="wells",
        style={"color": "#4FD2E8", "shape": "well", "size": 6, "labels": False},
        idField="Well", depthField="Total depth (ft)",
        source=("SBMM Monitoring Wells.xlsx, sheet WellConst (Groundwater "
                "Sampling). Coordinates NAD83(2011) CA SP Zone 2 ftUS as tabulated; "
                "elevations ft NAVD88."),
        points=pts), skipped


# --------------------------------------------------------------------------
def borings():
    p = os.path.join(SRC, "Geotechnical",
                     "Sulphur Bank Test Pit Soil Borings Location Coordinates20251210.xlsx")
    ws = openpyxl.load_workbook(p, data_only=True)["SB"]
    rows = list(ws.iter_rows(values_only=True))
    hdr = [txt(h) for h in rows[0]]
    col = {h: i for i, h in enumerate(hdr) if h}

    # "SB-6A" in the coordinate table is "SB06A" in the field records
    def key(loc):
        m = re.match(r"^SB-?(\d+)([A-Za-z]?)$", loc.strip())
        return "SB%02d%s" % (int(m.group(1)), m.group(2).upper()) if m else loc.upper()

    jp = os.path.join(SRC, "Geotechnical", "2025 Jacobs Investigation")
    sched = {}
    for r in openpyxl.load_workbook(
            os.path.join(jp, "Borings Schedule and Progress.xlsx"),
            data_only=True)["Sheet1"].iter_rows(min_row=6, values_only=True):
        if not r[2]:
            continue
        k = key(str(r[2]))
        d = num(r[4]) or num(r[3])          # actual drilled, else proposed
        if d is not None:
            prev = sched.get(k, (0, None))[0]
            sched[k] = (max(d, prev), txt(r[1]))

    waste = {}
    for r in openpyxl.load_workbook(
            os.path.join(jp, "Soil Profiles", "SBMM_Field Interpretted Waste Depth.xlsx"),
            data_only=True)["Sheet1"].iter_rows(min_row=5, values_only=True):
        if r[1]:
            waste[key(str(r[1]))] = (num(r[2]), txt(r[4]))

    pts, skipped = [], 0
    for r in rows[1:]:
        loc = txt(r[col["LOC_ID"]])
        x, y = num(r[col["E"]]), num(r[col["N"]])
        if not loc or not in_site(x, y):
            skipped += 1
            continue
        k = key(loc)
        depth, area = sched.get(k, (None, None))
        wd, wj = waste.get(k, (None, None))
        a = {
            "Waste area": area,
            "Ground elev (ft)": num(r[col["Elevation"]]),
            "Total depth (ft)": depth,
            "Interpreted waste depth (ft)": wd,
            "Waste depth basis": wj,
            "Latitude": num(r[col["Latitude"]]),
            "Longitude": num(r[col["Longtitude"]]),
        }
        pts.append({"id": loc, "x": round(x, 2), "y": round(y, 2),
                    "a": {k2: v for k2, v in a.items() if v is not None}})
    return dict(
        id="borings2025", name="Soil borings (2025 geotech)", kind="borings",
        style={"color": "#E8B34B", "shape": "boring", "size": 6, "labels": False},
        idField="Boring", depthField="Total depth (ft)",
        source=("Sulphur Bank Test Pit Soil Borings Location Coordinates20251210.xlsx, "
                "sheet SB (Geotechnical) for coordinates; depth drilled from 2025 Jacobs "
                "Investigation\\Borings Schedule and Progress.xlsx; interpreted waste "
                "depth from Soil Profiles\\SBMM_Field Interpretted Waste Depth.xlsx."),
        points=pts), skipped


def main():
    os.makedirs(OUT, exist_ok=True)
    for fn in (wells, borings):
        ds, skipped = fn()
        ds["crs"] = "EPSG:6418 (NAD83(2011) CA SP Zone 2, ftUS)"
        ds["baked"] = True
        dst = os.path.join(OUT, "ds_%s.json" % ds["id"])
        json.dump(ds, open(dst, "w"), indent=1)
        withd = sum(1 for p in ds["points"] if ds["depthField"] in p["a"])
        print("%-28s %3d points (%d skipped, %d with %s) -> %s"
              % (ds["name"], len(ds["points"]), skipped, withd, ds["depthField"],
                 os.path.basename(dst)))


if __name__ == "__main__":
    main()
